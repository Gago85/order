from flask import Flask, request, jsonify
from flask_cors import CORS
import csv
from datetime import datetime
import os
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)
from dotenv import load_dotenv
load_dotenv()

# 🔔 Функция отправки текста в Telegram
def send_telegram_text(chat_id, text):
    token = os.getenv("BOT_TOKEN")
    print("🔑 BOT_TOKEN =", token)
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text
    }
    requests.post(url, json=payload)


# 💰 Цены
prices = {
    "Стакан 0.1": 200.00,
    "Стакан 0.2": 10.00,
    "Стакан 0.3": 10.00,
    "Стакан 0.4": 10.00,
    "Стакан Шейк 0.4": 270.00,
    "Крышки 90": 200.00,
    "Крышки 80": 180.00,
    "Крышки Шейк": 100.00,
    "Трубочки": 350.00,
    "Мешалки": 300.00,
    "Ложки": 120.00,
    "Вилки": 120.00,
    "Лента Кассовая 80на80": 600.00,
    "Пакеты Майка": 50.00,
    "Мусорные 240л": 200.00,
    "Стекломой": 100.00,
    "Шуманит": 150.00,
    "Cafiza 1кг": 1200.00,
    "очистиель для форсунок 1л": 1200.00,
    "Фери": 150.00,
    "Белезна": 50.00,
    "пемалюкс": 100.00,
    "Ловушки": 350.00,
    "Губка": 50.00,
    "Фильтры для Bovilor 50шт": 400.00,
    "фильры для чая": 250.00,
    "Перчатки": 300.00,
    "Пергамент": 700.00,
    "Тряпка из микрофибры": 40.00,
    "Бумажные пакеты": 2000.00,
    "Салфетки": 10.00,
    "Зева": 200.00,
    "Тряпки половые": 165.00,
    "Фильтр для чая": 235.00,
    "Контейнера RK": 7.00,
    "Подставка 2х": 220.00,
    "Подставка 4х": 320.00,
    "Молоко 3,2%": 1000.00,
    "Сливки 11%": 200.00,
    "Кокосовый": 190.00,
    "Миндальный": 190.00,
    "Фундучный": 180.00,
    "Банановый": 180.00,
    "Соевый": 180.00,
    "Ягодный": 180.00,
    "Клюквеный": 180.00,
    "Облепиховый": 180.00,
    "Шиповник": 180.00,
    "Имбирный": 180.00,
    "маракуйя зм": 750.00,
    "Малина зм": 150.00,
    "ЭрлГрей": 250.00,
    "Сенча": 250.00,
    "Мята": 50.00,
    "Пряный Чай 250г": 950.00,
    "Апельсин св": 10.00,
    "Лимон": 10.00,
    "Лайм": 15.00,
    "БонАква 0.5": 700.00,
    "Татни 1 ящ": 1000.00,
    "Татни газ 1 ящ": 1000.00,
    "Апельсиновый сок 1л": 110.00,
    "Тоник 1 банка": 100.00,
    "Вода Газ 1,5л": 60.00,
    "Кофе эспрессо sidamo": 2050.00,
    "амаретто": 300.00,
    "Зерно фильтр": 850.00,
    "Арахисовая Паста": 250.00,
    "Какао 1кг": 1500.00,
    "Кедровые Орехи": 1200.00,
    "Матча": 800.00,
    "Гвоздика": 50.00,
    "Бадьян": 50.00,
    "цедра": 0.00,
    "Корица Пальчиковая": 50.00,
    "Корица Молотая": 50.00,
    "Сахар 1кг": 1500.00,
    "Солёная карамель": 300.00,
    "Карамель": 300.00,
    "Ваниль": 300.00,
    "Банан": 300.00,
    "Шоколад": 300.00,
    "Клубника": 300.00,
    "Hazel Nut": 300.00,
    "Фисташковый": 300.00,
    "Лаванда": 300.00,
    "Яблочный пирог": 300.00,
    "Кокос": 300.00,
    "Топпинг Банан": 200.00,
    "Топпинг Шоколад": 200.00,
    "Топпинг Карамель": 200.00,
    "Сгущенное молоко": 200.00,
    "Мороженое": 200.00
}

def load_prices_from_csv():
    filepath = os.path.join(os.path.dirname(__file__), "items.csv")
    loaded_prices = {}

    try:
        with open(filepath, newline="", encoding="utf-8-sig") as csv_file:
            reader = csv.DictReader(csv_file, delimiter=";")
            for row in reader:
                if row.get("active", "").strip().lower() == "false":
                    continue

                name = row.get("name", "").strip()
                price = row.get("price", "").strip().replace(",", ".")
                if not name:
                    continue

                try:
                    loaded_prices[name] = float(price or 0)
                except ValueError:
                    loaded_prices[name] = 0
    except FileNotFoundError:
        return {}

    return loaded_prices

prices.update(load_prices_from_csv())

# 📊 Генерация Excel накладной
def create_excel(data):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d %H:%M")
    filename = f"Приходная накладная - {date_str}.xlsx"
    filepath = os.path.join("/tmp", filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Накладная"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells("A1:D1")
    ws["A1"] = "GOODBUNS · Приходная накладная"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"] = "Документ сгенерирован автоматически через систему заказов GoodBuns"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = center

    ws["A4"] = "Дата:"
    ws["B4"] = date_str
    ws["A5"] = "Точка:"
    ws["B5"] = data.get("point", "")
    ws["A6"] = "День недели:"
    ws["B6"] = data.get("day", "")

    for row in range(4, 7):
        ws[f"A{row}"].font = bold

    start_row = 8
    headers = ["Наименование", "Количество", "Цена (₽)", "Сумма (₽)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    total = 0
    items = data.get("items", [])
    for i, item in enumerate(items):
        name = item["name"].strip()
        qty = int(item["qty"])
        price = prices.get(name, 0)
        amount = qty * price
        total += amount

        for j, val in enumerate([name, qty, price, amount], 1):
            cell = ws.cell(row=start_row + 1 + i, column=j)
            cell.value = val
            cell.border = thin_border
            if j in [2, 3, 4]:
                cell.alignment = center

    end_row = start_row + 1 + len(items)
    ws.cell(row=end_row, column=3).value = "ИТОГО"
    ws.cell(row=end_row, column=3).font = bold
    ws.cell(row=end_row, column=4).value = total
    ws.cell(row=end_row, column=4).font = bold
    ws.cell(row=end_row, column=4).alignment = center

    footer_row = end_row + 2
    ws.merge_cells(f"A{footer_row}:D{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = "Спасибо за заказ! Увидимся на смене ☕"
    footer.font = Font(italic=True, color="888888")
    footer.alignment = center

    wb.save(filepath)
    return filepath

# 📤 Отправка в Telegram
def send_to_telegram(filepath):
    bot_token = os.environ.get("BOT_TOKEN")
    chat_id = os.environ.get("CHAT_ID")
    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    with open(filepath, 'rb') as doc_file:
        files = {'document': doc_file}
        data = {'chat_id': chat_id}
        response = requests.post(url, files=files, data=data)
        print("📤 Ответ от Telegram (файл):", response.text)  # <--- ВАЖНО!
    
    return response.status_code == 200


# 📬 Обработка заказа
@app.route("/order", methods=["POST"])
def handle_order():
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Нет данных"}), 400
    try:
        filepath = create_excel(data)
        sent = send_to_telegram(filepath)
        return jsonify({"status": "ok", "saved": filepath, "telegram_sent": sent}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ☕️ Обработка Telegram webhook
pins = {
    "1111": "Магистраль Плаза",
    "2222": "Омега Плаза",
    "3333": "Черри Тауэр",
    "4444": "Гефест"
}

@app.route("/webhook", methods=["POST"])
def telegram_webhook():
    data = request.get_json()
    print("📩 Получено сообщение от Telegram:")
    print(data)

    if not data or "message" not in data:
        return jsonify({"status": "ignored"}), 200

    message = data["message"]
    chat_id = message["chat"]["id"]
    text = message.get("text", "").strip()

    if text == "/start":
        reply_markup = {
            "keyboard": [[
                {
                    "text": "📝 Оформить заказ",
                    "web_app": { "url": "https://gago85.github.io/order/" }
                }
            ]],
            "resize_keyboard": True
        }

        token = os.getenv("BOT_TOKEN")
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": "Нажми на кнопку ниже, чтобы оформить заказ 👇",
            "reply_markup": reply_markup
        }
        requests.post(url, json=payload)
        return jsonify({"status": "ok"}), 200

    elif text in pins:
        reply = f"✅ PIN принят. Добро пожаловать в {pins[text]}!"
    else:
        reply = "❌ Неверный PIN. Попробуйте снова."

    send_telegram_text(chat_id, reply)
    return jsonify({"status": "ok"}), 200
